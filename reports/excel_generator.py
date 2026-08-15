import os
import json
import logging
from datetime import datetime
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from data.models import StoreReportData

logger = logging.getLogger(__name__)

class ExcelGenerator:
    def __init__(self, output_dir: str):
        self.output_dir = output_dir
        os.makedirs(self.output_dir, exist_ok=True)

    def generate(self, data: StoreReportData) -> str:
        """
        Generates a professional Excel report with 4 sheets:
        1. Tong_Quan
        2. Chi_Tiet_Danh_Gia
        3. Khao_Sat_Ho_Tro
        4. Loi_Ton_Dong
        """
        wb = openpyxl.Workbook()
        # Remove default sheet
        default_sheet = wb.active
        wb.remove(default_sheet)

        # Style Definitions
        font_family = "Arial"
        navy_fill = PatternFill(start_color="0A2342", end_color="0A2342", fill_type="solid")
        zebra_fill = PatternFill(start_color="F2F5F8", end_color="F2F5F8", fill_type="solid")
        white_font = Font(name=font_family, size=11, bold=True, color="FFFFFF")
        
        title_font = Font(name=font_family, size=16, bold=True, color="0A2342")
        section_font = Font(name=font_family, size=12, bold=True, color="2C3E50")
        bold_font = Font(name=font_family, size=10, bold=True)
        regular_font = Font(name=font_family, size=10)
        italic_muted_font = Font(name=font_family, size=9, italic=True, color="7F8C8D")

        # Status fills
        green_fill = PatternFill(start_color="D4EDDA", end_color="D4EDDA", fill_type="solid") # light green
        green_font = Font(name=font_family, size=10, bold=True, color="155724")
        red_fill = PatternFill(start_color="F8D7DA", end_color="F8D7DA", fill_type="solid") # light red
        red_font = Font(name=font_family, size=10, bold=True, color="721C24")
        amber_fill = PatternFill(start_color="FFF3CD", end_color="FFF3CD", fill_type="solid") # light amber
        amber_font = Font(name=font_family, size=10, bold=True, color="856404")

        # Borders
        thin_side = Side(border_style="thin", color="D1D8E0")
        thin_border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)
        double_bottom = Border(bottom=Side(border_style="double", color="0A2342"), top=thin_side)

        # ----------------------------------------------------
        # SHEET 1: TỔNG QUAN
        # ----------------------------------------------------
        ws1 = wb.create_sheet(title="Tong_Quan")
        ws1.views.sheetView[0].showGridLines = True

        # Header Title
        ws1["A1"] = "BÁO CÁO ĐÁNH GIÁ CHẤT LƯỢNG CỬA HÀNG"
        ws1["A1"].font = title_font
        ws1.row_dimensions[1].height = 30

        report_date_str = data.form_response.report_date if data.form_response else datetime.now().strftime("%d/%m/%Y")
        ws1["A2"] = f"Hệ thống StoreVisit Pro - Ngày báo cáo: {report_date_str}"
        ws1["A2"].font = italic_muted_font

        # 1.1 Thông tin chung
        ws1["A4"] = "I. THÔNG TIN CHUNG"
        ws1["A4"].font = section_font
        
        meta_rows = [
            ("Mã cửa hàng", data.metadata.store_code),
            ("Tên cửa hàng", data.metadata.store_name),
            ("Địa chỉ", data.metadata.address),
            ("Khu vực / ASM", f"{data.metadata.region} / {data.metadata.asm_name}"),
            ("Cửa hàng trưởng", data.staff.cht_name),
            ("Thời gian đánh giá", f"{data.form_response.time_start if data.form_response else ''} - {data.form_response.time_end if data.form_response else ''}")
        ]
        
        curr_row = 5
        for label, val in meta_rows:
            ws1.cell(row=curr_row, column=1, value=label).font = bold_font
            ws1.cell(row=curr_row, column=2, value=val).font = regular_font
            ws1.cell(row=curr_row, column=1).border = thin_border
            ws1.cell(row=curr_row, column=2).border = thin_border
            curr_row += 1

        # 1.2 Hiệu quả hoạt động (MTD)
        curr_row += 1
        ws1.cell(row=curr_row, column=1, value="II. HIỆU QUẢ HOẠT ĐỘNG (MTD)").font = section_font
        curr_row += 1
        
        rev = data.revenue
        perf_rows = [
            ("Doanh thu thực tế (MTD)", rev.revenue_actual, "#,##0"),
            ("Chỉ tiêu doanh thu tháng", rev.revenue_target, "#,##0"),
            ("Tỷ lệ hoàn thành chỉ tiêu", rev.attainment_pct / 100.0, "0.0%"),
            ("Tỷ lệ tăng trưởng so tháng trước (MoM)", rev.mom_change_pct / 100.0, "+0.0%;-0.0%;0.0%")
        ]
        
        for label, val, num_fmt in perf_rows:
            c1 = ws1.cell(row=curr_row, column=1, value=label)
            c2 = ws1.cell(row=curr_row, column=2, value=val)
            c1.font = bold_font
            c2.font = regular_font
            c2.number_format = num_fmt
            c1.border = thin_border
            c2.border = thin_border
            curr_row += 1

        # 1.3 Xếp loại hạng mục
        curr_row += 1
        ws1.cell(row=curr_row, column=1, value="III. XẾP LOẠI HẠNG MỤC VẬN HÀNH").font = section_font
        curr_row += 1
        
        headers_ws1 = ["Hạng mục đánh giá", "Kết quả xếp loại"]
        for col_idx, text in enumerate(headers_ws1, 1):
            cell = ws1.cell(row=curr_row, column=col_idx, value=text)
            cell.font = white_font
            cell.fill = navy_fill
            cell.alignment = Alignment(horizontal="center")
        ws1.row_dimensions[curr_row].height = 24
        curr_row += 1
        
        categories = [
            ("1. Mặt tiền cửa hàng (Frontage)", data.frontage_rating or "Đạt"),
            ("2. Không gian bên trong (Inner Space)", data.form_response.rating_inner if data.form_response else "Đạt"),
            ("3. Trưng bày hàng hóa (Merchandising)", data.form_response.rating_merch if data.form_response else "Đạt"),
            ("4. Nhân sự & Tác phong (Staff)", data.form_response.rating_staff if data.form_response else "Đạt"),
            ("5. Cơ sở vật chất & An ninh (CSVC)", data.form_response.rating_csvc if data.form_response else "Đạt")
        ]
        
        for cat, rating in categories:
            c1 = ws1.cell(row=curr_row, column=1, value=cat)
            c2 = ws1.cell(row=curr_row, column=2, value=rating)
            c1.font = regular_font
            c2.font = bold_font
            c2.alignment = Alignment(horizontal="center")
            c1.border = thin_border
            c2.border = thin_border
            
            # Apply styling according to rating
            rl = rating.lower()
            if "chưa đạt" in rl or "không đạt" in rl:
                c2.fill = red_fill
                c2.font = red_font
            elif rl == "tốt":
                c2.fill = green_fill
                c2.font = green_font
            else:
                c2.fill = amber_fill
                c2.font = amber_font
            curr_row += 1

        # ----------------------------------------------------
        # SHEET 2: CHI TIẾT ĐÁNH GIÁ (ALL CHECKLIST ITEMS)
        # ----------------------------------------------------
        ws2 = wb.create_sheet(title="Chi_Tiet_Danh_Gia")
        ws2.views.sheetView[0].showGridLines = True

        ws2["A1"] = "CHI TIẾT ĐÁNH GIÁ TIÊU CHÍ VẬN HÀNH"
        ws2["A1"].font = title_font
        ws2.row_dimensions[1].height = 30
        
        ws2["A2"] = f"Cửa hàng: {data.metadata.store_name} ({data.metadata.store_code})"
        ws2["A2"].font = italic_muted_font
        
        headers_ws2 = ["Mã", "Hạng mục", "Nội dung tiêu chí", "Đánh giá", "Ghi chú lỗi / Chi tiết", "Ảnh chụp hiện trạng (Drive)"]
        ws2_row = 4
        for col_idx, text in enumerate(headers_ws2, 1):
            cell = ws2.cell(row=ws2_row, column=col_idx, value=text)
            cell.font = white_font
            cell.fill = navy_fill
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws2.row_dimensions[ws2_row].height = 26
        ws2_row += 1

        # Mappings of checklist keys to friendly names
        section_names = {
            "frontage": "Mặt tiền",
            "inner": "Không không gian bên trong",
            "merch_ap": "Trưng bày An Phước",
            "merch_pie": "Trưng bày Pierre Cardin",
            "merch_ab": "Trưng bày Anamai/Bonjour",
            "merch_anamai": "Trưng bày Anamai",
            "merch_bonjour": "Trưng bày Bonjour",
            "merch_pk": "Phụ kiện",
            "warehouse": "Kho/Phòng thử",
            "stockroom": "Kho hàng",
            "fitting_room": "Phòng thử đồ",
            "toilet": "Nhà vệ sinh",
            "fire_safety": "PCCC & Thoát hiểm",
            "cashier": "Thu ngân",
            "packaging_security": "Bao bì & An ninh",
            "staff": "Nhân sự & Tác phong",
            "security_guard": "Bảo vệ"
        }

        # Retrieve photos drive map from form response to link in grid
        photo_url_map = {}
        if data.form_response:
            for photo in data.form_response.photos:
                # Key on section + index
                photo_url_map[f"{photo.section}_{photo.index}"] = photo.drive_url

        if data.form_response and data.form_response.checklist_json:
            try:
                c_data = json.loads(data.form_response.checklist_json)
                sections = c_data.get("sections", {})
                
                # Sắp xếp các section để duyệt tuần tự
                for sec_key, sec_val in sections.items():
                    sec_name = section_names.get(sec_key, sec_key.upper())
                    items = sec_val.get("items", [])
                    
                    for item in items:
                        c_id = item.get("id", "")
                        c_lbl = item.get("label", "")
                        c_eval = item.get("eval", "Đạt")
                        c_note = item.get("note", "")
                        
                        # Find photo drive url
                        photo_urls = []
                        # Look for photo before
                        raw_before = item.get("photo_before", "")
                        if raw_before:
                            photo_urls.append(f"Trước: https://drive.google.com/open?id={raw_before}")
                        # Look for photo after
                        raw_after = item.get("photo_after", "")
                        if raw_after:
                            photo_urls.append(f"Sau: https://drive.google.com/open?id={raw_after}")
                            
                        photo_link_str = "\n".join(photo_urls) if photo_urls else "-"
                        
                        # Populate row
                        ws2.cell(row=ws2_row, column=1, value=c_id).font = bold_font
                        ws2.cell(row=ws2_row, column=2, value=sec_name).font = regular_font
                        ws2.cell(row=ws2_row, column=3, value=c_lbl).font = regular_font
                        
                        cell_eval = ws2.cell(row=ws2_row, column=4, value=c_eval)
                        cell_eval.font = bold_font
                        cell_eval.alignment = Alignment(horizontal="center")
                        
                        # Style eval
                        ev = c_eval.lower()
                        if "chưa đạt" in ev or "không đạt" in ev:
                            cell_eval.fill = red_fill
                            cell_eval.font = red_font
                        elif ev == "tốt":
                            cell_eval.fill = green_fill
                            cell_eval.font = green_font
                        else:
                            cell_eval.fill = amber_fill
                            cell_eval.font = amber_font

                        ws2.cell(row=ws2_row, column=5, value=c_note or "-").font = regular_font
                        
                        cell_photo = ws2.cell(row=ws2_row, column=6, value=photo_link_str)
                        cell_photo.font = italic_muted_font
                        cell_photo.alignment = Alignment(wrap_text=True)
                        
                        # Borders
                        for col in range(1, 7):
                            ws2.cell(row=ws2_row, column=col).border = thin_border
                            
                        # Set height to accommodate wrapped text
                        ws2.row_dimensions[ws2_row].height = 20
                        ws2_row += 1
            except Exception as e:
                logger.error(f"Error expanding checklist in Excel: {e}")
                ws2.cell(row=ws2_row, column=1, value="Lỗi tải chi tiết checklist JSON").font = red_font
        else:
            # Fallback: No checklist json, just print overall ratings
            ws2.cell(row=ws2_row, column=1, value="N/A").font = bold_font
            ws2.cell(row=ws2_row, column=2, value="Thông tin").font = regular_font
            ws2.cell(row=ws2_row, column=3, value="Dữ liệu cũ chưa có checklist_json").font = regular_font
            ws2.cell(row=ws2_row, column=4, value="N/A").font = regular_font
            ws2.cell(row=ws2_row, column=5, value="Vui lòng cập nhật form mới để có bảng chi tiết.").font = italic_muted_font
            for col in range(1, 7):
                ws2.cell(row=ws2_row, column=col).border = thin_border
            ws2_row += 1

        # ----------------------------------------------------
        # SHEET 3: KHẢO SÁT HỖ TRỢ (SURVEY A1 -> E7)
        # ----------------------------------------------------
        ws3 = wb.create_sheet(title="Khao_Sat_Ho_Tro")
        ws3.views.sheetView[0].showGridLines = True

        ws3["A1"] = "GHI NHẬN KHẢO SÁT & ĐỀ XUẤT HỖ TRỢ CỬA HÀNG"
        ws3["A1"].font = title_font
        ws3.row_dimensions[1].height = 30
        
        ws3["A2"] = f"ASM: {data.metadata.asm_name} - Cửa hàng: {data.metadata.store_name}"
        ws3["A2"].font = italic_muted_font
        
        headers_ws3 = ["Mã câu hỏi", "Hạng mục khảo sát", "Nội dung khảo sát chi tiết", "Ý kiến ghi nhận thực tế"]
        ws3_row = 4
        for col_idx, text in enumerate(headers_ws3, 1):
            cell = ws3.cell(row=ws3_row, column=col_idx, value=text)
            cell.font = white_font
            cell.fill = navy_fill
            cell.alignment = Alignment(horizontal="center")
        ws3.row_dimensions[ws3_row].height = 26
        ws3_row += 1

        survey_questions = [
            ("A1", "A. NHÂN SỰ CỬA HÀNG", "Đủ định biên nhân sự vận hành"),
            ("A2", "A. NHÂN SỰ CỬA HÀNG", "Tình trạng nhân viên nghỉ việc"),
            ("A3", "A. NHÂN SỰ CỬA HÀNG", "Lý do nghỉ việc chính"),
            ("A4", "A. NHÂN SỰ CỬA HÀNG", "Lịch phân ca & bố trí làm việc"),
            ("A5", "A. NHÂN SỰ CỬA HÀNG", "Đào tạo nghiệp vụ & sản phẩm mới"),
            
            ("B1", "B. HÀNG HÓA & TRƯNG BÀY", "Đầy đủ hàng hóa/mã hàng (Nguyên giá, Sale)"),
            ("B2", "B. HÀNG HÓA & TRƯNG BÀY", "Tiến trình xử lý tồn lâu, hàng bán chậm"),
            ("B3", "B. HÀNG HÓA & TRƯNG BÀY", "Thực hiện VM tiêu chuẩn trưng bày"),
            ("B4", "B. HÀNG HÓA & TRƯNG BÀY", "Quy trình xử lý lỗi/sự cố hàng lỗi dơ"),
            
            ("C1", "C. Ý KIẾN CỬA HÀNG", "Lắng nghe ý kiến phản hồi từ nhân viên"),
            ("C2", "C. Ý KIẾN CỬA HÀNG", "Mong muốn lớn nhất đối với quản lý"),
            ("C3", "C. Ý KIẾN CỬA HÀNG", "Khó khăn gặp phải chưa giải quyết được"),
            ("C4", "C. Ý KIẾN CỬA HÀNG", "Ý kiến đóng góp tự do khác"),
            
            ("D1", "D. VẬN HÀNH NỘI BỘ", "Quy trình đóng / mở cửa hàng chuẩn"),
            ("D2", "D. VẬN HÀNH NỘI BỘ", "Công tác kiểm kho định kỳ hàng tháng"),
            ("D3", "D. VẬN HÀNH NỘI BỘ", "Gửi báo cáo doanh số, ca trực đúng hạn"),
            ("D4", "D. VẬN HÀNH NỘI BỘ", "Quy trình xử lý sự cố khẩn cấp (mất điện, trộm...)"),
            ("D5", "D. VẬN HÀNH NỘI BỘ", "Hiệu quả sử dụng phần mềm, POS, két sắt"),
            ("D6", "D. VẬN HÀNH NỘI BỘ", "Nhu cầu đào tạo kỹ năng bổ sung"),
            
            ("E1", "E. HIỆU QUẢ BÁN HÀNG", "Lý do khách rời đi không mua hàng"),
            ("E2", "E. HIỆU QUẢ BÁN HÀNG", "Độ tự tin tư vấn chất liệu, phối đồ của NV"),
            ("E3", "E. HIỆU QUẢ BÁN HÀNG", "Áp dụng kỹ thuật bán thêm (Upsell/Cross-sell)"),
            ("E4", "E. HIỆU QUẢ BÁN HÀNG", "Động lực làm việc từ chính sách thưởng/hoa hồng"),
            ("E5", "E. HIỆU QUẢ BÁN HÀNG", "Đánh giá doanh số MTD so với cùng kỳ"),
            ("E6", "E. HIỆU QUẢ BÁN HÀNG", "Mở rộng tệp khách hàng tiềm năng"),
            ("E7", "E. HIỆU QUẢ BÁN HÀNG", "Khảo sát hợp đồng bán lẻ trong bán kính 2km (cơ quan, trường học...)"),
        ]

        survey_data = {}
        if data.form_response and data.form_response.checklist_json:
            try:
                c_data = json.loads(data.form_response.checklist_json)
                survey_data = c_data.get("survey", {})
            except Exception as e:
                logger.error(f"Error loading survey from checklist json in excel gen: {e}")

        for code, category, question in survey_questions:
            ans_obj = survey_data.get(code, {})
            ans_text = ans_obj.get("answer", "") if isinstance(ans_obj, dict) else str(ans_obj)
            if not ans_text or ans_text == "undefined":
                ans_text = "-"
                
            ws3.cell(row=ws3_row, column=1, value=code).font = bold_font
            ws3.cell(row=ws3_row, column=1).alignment = Alignment(horizontal="center")
            ws3.cell(row=ws3_row, column=2, value=category).font = bold_font
            ws3.cell(row=ws3_row, column=3, value=question).font = regular_font
            
            c_ans = ws3.cell(row=ws3_row, column=4, value=ans_text)
            c_ans.font = regular_font
            c_ans.alignment = Alignment(wrap_text=True)
            
            # Apply light zebra to groups
            if code[0] in ["B", "D"]:
                for col in range(1, 5):
                    ws3.cell(row=ws3_row, column=col).fill = zebra_fill
            
            for col in range(1, 5):
                ws3.cell(row=ws3_row, column=col).border = thin_border
            ws3_row += 1

        # ----------------------------------------------------
        # SHEET 4: LỖI TỒN ĐỌNG (OPERATIONAL ISSUES)
        # ----------------------------------------------------
        ws4 = wb.create_sheet(title="Loi_Ton_Dong")
        ws4.views.sheetView[0].showGridLines = True

        ws4["A1"] = "DANH SÁCH LỖI VẬN HÀNH & KẾ HOẠCH KHẮC PHỤC"
        ws4["A1"].font = title_font
        ws4.row_dimensions[1].height = 30
        
        ws4["A2"] = f"Cửa hàng: {data.metadata.store_name} - Số lỗi hiện tại: {len(data.issues)}"
        ws4["A2"].font = italic_muted_font

        headers_ws4 = ["STT", "Hạng mục lỗi", "Chi tiết lỗi ghi nhận", "Người chịu trách nhiệm", "Hạn khắc phục / Kế hoạch", "Trạng thái"]
        ws4_row = 4
        for col_idx, text in enumerate(headers_ws4, 1):
            cell = ws4.cell(row=ws4_row, column=col_idx, value=text)
            cell.font = white_font
            cell.fill = navy_fill
            cell.alignment = Alignment(horizontal="center")
        ws4.row_dimensions[ws4_row].height = 26
        ws4_row += 1

        if not data.issues:
            ws4.cell(row=ws4_row, column=1, value="Chúc mừng! Không ghi nhận lỗi tồn đọng tại cửa hàng.").font = green_font
            ws4.merge_cells(start_row=ws4_row, start_column=1, end_row=ws4_row, end_column=6)
            ws4.row_dimensions[ws4_row].height = 24
            for col in range(1, 7):
                ws4.cell(row=ws4_row, column=col).border = thin_border
        else:
            for idx, issue in enumerate(data.issues, 1):
                ws4.cell(row=ws4_row, column=1, value=idx).font = bold_font
                ws4.cell(row=ws4_row, column=1).alignment = Alignment(horizontal="center")
                
                ws4.cell(row=ws4_row, column=2, value=issue.label).font = bold_font
                ws4.cell(row=ws4_row, column=3, value=issue.issue).font = regular_font
                ws4.cell(row=ws4_row, column=4, value=issue.assignee).font = regular_font
                ws4.cell(row=ws4_row, column=5, value=issue.notes).font = regular_font
                
                cell_status = ws4.cell(row=ws4_row, column=6, value=issue.status)
                cell_status.font = bold_font
                cell_status.alignment = Alignment(horizontal="center")
                
                # Status formatting
                st = issue.status.lower()
                if "chưa" in st or "pending" in st:
                    cell_status.fill = red_fill
                    cell_status.font = red_font
                elif "đã" in st or "resolved" in st or "done" in st:
                    cell_status.fill = green_fill
                    cell_status.font = green_font
                else:
                    cell_status.fill = amber_fill
                    cell_status.font = amber_font
                
                for col in range(1, 7):
                    ws4.cell(row=ws4_row, column=col).border = thin_border
                ws4_row += 1

        # Auto-adjust column width for all sheets
        for sheet in wb.worksheets:
            for col in sheet.columns:
                max_len = 0
                for cell in col:
                    if cell.value:
                        # Avoid huge title cell expansion
                        if cell.coordinate in ["A1", "A2"]:
                            continue
                        lines = str(cell.value).split("\n")
                        for line in lines:
                            max_len = max(max_len, len(line))
                col_letter = get_column_letter(col[0].column)
                # Set reasonable width
                sheet.column_dimensions[col_letter].width = max(max_len + 3, 12)

        # Explicit tweaks to column sizes for readability
        ws2.column_dimensions["C"].width = 40 # Tiêu chí
        ws2.column_dimensions["E"].width = 35 # Ghi chú
        ws2.column_dimensions["F"].width = 30 # Link ảnh
        
        ws3.column_dimensions["C"].width = 45 # Nội dung khảo sát
        ws3.column_dimensions["D"].width = 40 # Câu trả lời

        ws4.column_dimensions["C"].width = 45 # Chi tiết lỗi
        ws4.column_dimensions["E"].width = 35 # Kế hoạch

        # Save workbook
        from utils.filename_formatter import format_store_output_filename
        r_date = data.form_response.report_date if data.form_response else ""
        filename = format_store_output_filename(data.metadata.store_name, data.metadata.asm_name, r_date, "xlsx")
        dest_path = os.path.join(self.output_dir, filename)
        wb.save(dest_path)
        logger.info(f"Generated excel report successfully at: {dest_path}")
        return dest_path
