import os
import pandas as pd
from typing import List
from data.models import MarketSurveyResponse

class SurveyConsolidator:
    def consolidate_to_excel(self, responses: List[MarketSurveyResponse], output_path: str):
        if not responses:
            print("No responses to consolidate.")
            df_empty = pd.DataFrame([{"Thông báo": "Không có phản hồi khảo sát nào được phê duyệt"}])
            df_empty.to_excel(output_path, index=False)
            return
            
        # 1. Sheet: Detailed Responses
        rows = []
        for r in responses:
            row = {
                "Mã phản hồi": r.response_id,
                "Mã cửa hàng": r.store_code,
                "Khu vực/Cụm": r.region,
                "QLKD phụ trách": r.qlkd_asm,
                "Người trả lời": r.respondent_name,
                "Chức danh": r.respondent_role,
                "Số người thảo luận": r.discussion_count,
                "Ngày khảo sát": r.survey_date,
                "Thay đổi của khách hàng": r.customer_change,
                "Nhu cầu tăng": ", ".join(r.demand_increase),
                "Nguyên nhân mất cơ hội bán": ", ".join(r.lost_sale_reasons),
                "Nguyên nhân lớn nhất": r.lost_sale_top1,
                "Sản phẩm cần bổ sung": ", ".join(r.product_gap),
                "Khoảng giá chấp nhận": r.acceptable_price,
                "Nhóm cần hỗ trợ": ", ".join(r.support_categories),
                "Giải pháp đề xuất": r.suggested_solution,
                "Mùa vụ / Cơ hội địa phương": r.local_opportunity,
                "Thời hạn hàng cần có": r.need_before_date,
                "Kiến nghị ưu tiên": r.store_recommendation,
                "Trạng thái QC": r.qc_status
            }
            rows.append(row)
            
        df_details = pd.DataFrame(rows)
        
        # 2. Sheet: Lost Sales Summary
        all_reasons = []
        for r in responses:
            all_reasons.extend(r.lost_sale_reasons)
        df_reasons = pd.Series(all_reasons).value_counts().reset_index()
        df_reasons.columns = ["Nguyên nhân mất cơ hội bán", "Số cửa hàng ghi nhận"]
        
        top1_reasons = [r.lost_sale_top1 for r in responses if r.lost_sale_top1]
        df_top1 = pd.Series(top1_reasons).value_counts().reset_index()
        df_top1.columns = ["Nguyên nhân lớn nhất (Top 1)", "Số cửa hàng chọn"]
        
        # 3. Sheet: Product Gaps Summary
        all_gaps = []
        for r in responses:
            all_gaps.extend(r.product_gap)
        df_gaps = pd.Series(all_gaps).value_counts().reset_index()
        df_gaps.columns = ["Nhóm sản phẩm cần bổ sung", "Số cửa hàng đề xuất"]
        
        prices = [r.acceptable_price for r in responses if r.acceptable_price]
        df_prices = pd.Series(prices).value_counts().reset_index()
        df_prices.columns = ["Khoảng giá khách chấp nhận", "Số cửa hàng chọn"]
        
        # 4. Sheet: Support Requests
        support_rows = []
        for r in responses:
            for cat in r.support_categories:
                support_rows.append({
                    "Cửa hàng": r.store_code,
                    "Khu vực": r.region,
                    "QLKD": r.qlkd_asm,
                    "Nhóm hỗ trợ": cat,
                    "Đề xuất giải pháp": r.suggested_solution
                })
        df_support = pd.DataFrame(support_rows) if support_rows else pd.DataFrame(columns=["Cửa hàng", "Khu vực", "QLKD", "Nhóm hỗ trợ", "Đề xuất giải pháp"])
        
        os.makedirs(os.path.dirname(output_path), exist_ok=True)
        
        with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
            df_details.to_excel(writer, sheet_name="Phản hồi chi tiết", index=False)
            df_reasons.to_excel(writer, sheet_name="Phân tích mất cơ hội bán", index=False, startrow=0)
            df_top1.to_excel(writer, sheet_name="Phân tích mất cơ hội bán", index=False, startrow=len(df_reasons) + 4)
            df_gaps.to_excel(writer, sheet_name="Nhu cầu sản phẩm mới", index=False, startrow=0)
            df_prices.to_excel(writer, sheet_name="Nhu cầu sản phẩm mới", index=False, startrow=len(df_gaps) + 4)
            df_support.to_excel(writer, sheet_name="Đề xuất hỗ trợ", index=False)
            
            wb = writer.book
            
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            from openpyxl.utils import get_column_letter
            
            header_font = Font(name="Segoe UI", size=11, bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="0A2342", end_color="0A2342", fill_type="solid")
            cell_font = Font(name="Segoe UI", size=10)
            thin_border = Border(
                left=Side(style="thin", color="D3D3D3"),
                right=Side(style="thin", color="D3D3D3"),
                top=Side(style="thin", color="D3D3D3"),
                bottom=Side(style="thin", color="D3D3D3")
            )
            
            for name in wb.sheetnames:
                ws = wb[name]
                for col in ws.columns:
                    max_len = 0
                    col_letter = get_column_letter(col[0].column)
                    for cell in col:
                        cell.font = cell_font
                        cell.border = thin_border
                        if cell.value:
                            max_len = max(max_len, len(str(cell.value)))
                    ws.column_dimensions[col_letter].width = max(max_len + 3, 12)
                    
                if name == "Phân tích mất cơ hội bán":
                    for col_idx in range(1, len(df_reasons.columns) + 1):
                        cell = ws.cell(row=1, column=col_idx)
                        cell.font = header_font
                        cell.fill = header_fill
                        cell.alignment = Alignment(horizontal="center")
                    start_row_2 = len(df_reasons) + 5
                    for col_idx in range(1, len(df_top1.columns) + 1):
                        cell = ws.cell(row=start_row_2, column=col_idx)
                        cell.font = header_font
                        cell.fill = header_fill
                        cell.alignment = Alignment(horizontal="center")
                elif name == "Nhu cầu sản phẩm mới":
                    for col_idx in range(1, len(df_gaps.columns) + 1):
                        cell = ws.cell(row=1, column=col_idx)
                        cell.font = header_font
                        cell.fill = header_fill
                        cell.alignment = Alignment(horizontal="center")
                    start_row_2 = len(df_gaps) + 5
                    for col_idx in range(1, len(df_prices.columns) + 1):
                        cell = ws.cell(row=start_row_2, column=col_idx)
                        cell.font = header_font
                        cell.fill = header_fill
                        cell.alignment = Alignment(horizontal="center")
                else:
                    for col_idx in range(1, ws.max_column + 1):
                        cell = ws.cell(row=1, column=col_idx)
                        cell.font = header_font
                        cell.fill = header_fill
                        cell.alignment = Alignment(horizontal="center")
                        
        print(f"Excel report consolidated successfully to: {output_path}")
