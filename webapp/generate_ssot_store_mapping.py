# -*- coding: utf-8 -*-
"""
StoreVisit/webapp/generate_ssot_store_mapping.py
═══════════════════════════════════════════════════════════════
ĐỒNG BỘ 185 CỬA HÀNG & 9 QLKD CHUẨN SSOT TỪ StoresInfo.xlsx
═══════════════════════════════════════════════════════════════
"""
import openpyxl
import json
from pathlib import Path

SRC_XLSX = Path(r"C:\All_Report\1_Mapping\StoresInfo.xlsx")
OUT_JSON = Path(r"C:\All_Report\8_RETAIL_COMMANDER\StoreVisit\webapp\store_mapping.json")

def generate_mapping():
    wb = openpyxl.load_workbook(SRC_XLSX, data_only=True)
    ws = wb.active
    
    headers = [str(ws.cell(1, c).value or '').strip() for c in range(1, ws.max_column + 1)]
    c_code = headers.index('STORECODE') + 1
    c_name = headers.index('TÊN CỬA HÀNG - CHUẨN') + 1
    c_asm  = headers.index('ASM') + 1
    c_reg  = headers.index('VÙNG') + 1 if 'VÙNG' in headers else headers.index('REGION') + 1
    
    asms_set = set()
    regions_set = set()
    mapping_by_asm = {}
    mapping_by_region = {}
    
    total_stores = 0
    for r in range(2, ws.max_row + 1):
        code = str(ws.cell(r, c_code).value or '').strip()
        if not code or code == '0':
            continue
            
        name = str(ws.cell(r, c_name).value or '').strip()
        asm = str(ws.cell(r, c_asm).value or '').strip()
        reg = str(ws.cell(r, c_reg).value or '').strip()
        
        if not asm:
            asm = 'Khác'
        if not reg:
            reg = 'Khác'
            
        asms_set.add(asm)
        regions_set.add(reg)
        
        store_entry = {
            'store_code': code,
            'store_name': name,
            'asm': asm,
            'region': reg
        }
        
        mapping_by_asm.setdefault(asm, []).append(store_entry)
        mapping_by_region.setdefault(reg, []).append(store_entry)
        total_stores += 1
        
    out_data = {
        'total_stores': total_stores,
        'asms': sorted(list(asms_set)),
        'regions': sorted(list(regions_set)),
        'mapping_by_asm': mapping_by_asm,
        'mapping_by_region': mapping_by_region
    }
    
    OUT_JSON.write_text(json.dumps(out_data, ensure_ascii=False, indent=2), encoding='utf-8')
    print(f"✅ Đã đồng bộ {total_stores} cửa hàng & {len(asms_set)} ASM vào {OUT_JSON}")
    print("  Danh sách ASMs:", sorted(list(asms_set)))

if __name__ == '__main__':
    generate_mapping()
