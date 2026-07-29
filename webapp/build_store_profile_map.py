# -*- coding: utf-8 -*-
"""
build_store_profile_map.py — Sinh STORE_PROFILE_MAP cho Code.gs
================================================================
Đọc StoresInfo.xlsx → xuất map: STORECODE → hồ sơ cửa hàng phục vụ
render checklist động (thương hiệu áp dụng + có/không có bảo vệ + bỏ khảo sát).

Nguồn: C:\\All_Report\\1_Mapping\\StoresInfo.xlsx
  Cột K BRAND     : AP-PIE-A-B / AP-PIE / A-B / 0   (AP=An Phước, PIE=Pierre Cardin, A=Anamai, B=Bonjour)
  Cột L STORECODE : mã dùng trong form
  Cột N STORETYPE : Standalone Boutique / Concession / Maison / Flagship Store / Mono-brand Store / Online Store

Quy tắc (đã chốt với user):
  - has_guard = storetype ∈ {Standalone Boutique, Maison, Flagship Store, Mono-brand Store}
                AND brand có AP hoặc PIE (KHÔNG phải CH chỉ Anamai-Bonjour)
  - Concession (trong mall) & Online → không bảo vệ
  - Online Store hoặc brand '0' → skip_survey = True

Output:
  - webapp/store_profile_map.js   (var STORE_PROFILE_MAP = {...};  — nhúng vào Code.gs)
  - webapp/store_profile_map.json (cho Python report dùng nếu cần)

Chạy lại mỗi khi StoresInfo.xlsx đổi:
  .venv\\Scripts\\python.exe webapp\\build_store_profile_map.py
"""
import os
import json
import openpyxl

SRC = r"C:\All_Report\1_Mapping\StoresInfo.xlsx"
OUT_DIR = os.path.dirname(os.path.abspath(__file__))

GUARD_TYPES = {"Standalone Boutique", "Maison", "Flagship Store", "Mono-brand Store"}
NO_SURVEY_TYPES = {"Online Store"}


def brand_tokens(brand: str):
    b = (brand or "").strip().upper()
    if not b or b == "0":
        return set()
    return set(t.strip() for t in b.split("-") if t.strip())


def build():
    wb = openpyxl.load_workbook(SRC, data_only=True)
    ws = wb.active
    header = [str(ws.cell(row=1, column=c).value or "").strip().upper()
              for c in range(1, ws.max_column + 1)]

    def col(*names):
        for nm in names:
            if nm in header:
                return header.index(nm) + 1
        return None

    c_brand = col("BRAND")
    c_code = col("STORECODE", "CODE")
    c_type = col("STORETYPE", "STORE TYPE")
    c_name = col("TÊN CỬA HÀNG - CHUẨN", "CỬA HÀNG")

    profiles = {}
    for r in range(2, ws.max_row + 1):
        code = str(ws.cell(row=r, column=c_code).value or "").strip()
        if not code:
            continue
        brand = str(ws.cell(row=r, column=c_brand).value or "").strip()
        stype = str(ws.cell(row=r, column=c_type).value or "").strip()
        tokens = brand_tokens(brand)

        has_ap = "AP" in tokens
        has_pie = "PIE" in tokens
        has_anamai = "A" in tokens
        has_bonjour = "B" in tokens
        is_ab_only = (has_anamai or has_bonjour) and not (has_ap or has_pie)

        skip_survey = (stype in NO_SURVEY_TYPES) or (not tokens)
        has_guard = (stype in GUARD_TYPES) and (has_ap or has_pie) and not is_ab_only

        profiles[code] = {
            "brand": brand,
            "storetype": stype,
            "has_ap": has_ap,
            "has_pie": has_pie,
            "has_anamai": has_anamai,
            "has_bonjour": has_bonjour,
            "is_ab_only": is_ab_only,
            "has_guard": has_guard,
            "skip_survey": skip_survey,
        }

    # JSON
    json_path = os.path.join(OUT_DIR, "store_profile_map.json")
    with open(json_path, "w", encoding="utf-8") as f:
        json.dump(profiles, f, ensure_ascii=False, indent=2)

    # JS snippet cho Code.gs
    js_path = os.path.join(OUT_DIR, "store_profile_map.js")
    with open(js_path, "w", encoding="utf-8") as f:
        f.write("// AUTO-GENERATED bởi build_store_profile_map.py — KHÔNG sửa tay.\n")
        f.write("// Nguồn: StoresInfo.xlsx (cột BRAND, STORECODE, STORETYPE).\n")
        f.write("var STORE_PROFILE_MAP = ")
        f.write(json.dumps(profiles, ensure_ascii=False, indent=2))
        f.write(";\n")

    # Summary
    total = len(profiles)
    n_guard = sum(1 for p in profiles.values() if p["has_guard"])
    n_skip = sum(1 for p in profiles.values() if p["skip_survey"])
    n_abonly = sum(1 for p in profiles.values() if p["is_ab_only"])
    print(f"Stores: {total} | có bảo vệ: {n_guard} | bỏ khảo sát: {n_skip} | chỉ A-B: {n_abonly}")
    print(f"Wrote: {js_path}")
    print(f"Wrote: {json_path}")
    # Sample
    for code in list(profiles)[:3]:
        print(" ", code, profiles[code])


if __name__ == "__main__":
    build()
